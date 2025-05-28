from flask import Flask, render_template, request, redirect, url_for, session, flash
from flask_sqlalchemy import SQLAlchemy
import sqlite3
from werkzeug.security import check_password_hash
from flask_sqlalchemy import SQLAlchemy
import csv
from io import StringIO
from flask import Response
import pandas as pd
from flask import send_file
from io import BytesIO
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
from datetime import datetime
import locale 
from flask import g
import sqlite3
import io
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from collections import Counter
from collections import defaultdict






app = Flask(__name__)
app.secret_key = 'gizli_anahtar'
DB_PATH = 'rehber.db'

def create_table():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("DROP TABLE IF EXISTS siparisler")
    cursor.execute("""
        CREATE TABLE siparisler (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            kalan_mktar REAL,
            sevk_miktar REAL,
            sipno TEXT,
            musteri TEXT,
            termin TEXT,
            hafta TEXT,
            artno TEXT,
            ebat TEXT,
            renkno TEXT,
            renk TEXT,
            gr_m2 REAL,
            bordur TEXT,
            nakis TEXT,
            baski TEXT,
            sip_ad REAL,
            faz_yk_ad TEXT,
            etiket TEXT,
            poset TEXT,
            pos_ici TEXT,
            koli TEXT,
            koli_ici TEXT,
            net_ks REAL,
            adetli_net REAL,
            topl_net_koli REAL,
            max_ks REAL,
            adetli_max REAL,
            faz_topl_koli REAL,
            netm3 REAL,
            maxm3 REAL,
            ort_m3 REAL,
            dokuma_mkt REAL,
            bc_ad TEXT,
            bg_ad TEXT,
            mk_ad TEXT,
            boy_dikim REAL,
            en_dikim REAL,
            fas_cik REAL,
            fas_gir REAL,
            nak_cik REAL,
            nak_gir REAL,
            klt1 REAL,
            klt2 REAL,
            klt3 REAL,
            order_no TEXT
        )
    """)
    conn.commit()
    conn.close()

def get_db_connection():
    conn = sqlite3.connect('rehber.db')
    conn.row_factory = sqlite3.Row
    return conn

@app.route('/')
def index():
    conn = get_db_connection()
    kisiler = conn.execute('SELECT * FROM kisiler').fetchall()
    conn.close()
    return render_template_tarihli('base.html', rehber=kisiler)


def tarih_formatla(tarih):
    if isinstance(tarih, datetime):
        return tarih.strftime('%d.%m.%Y')
    elif isinstance(tarih, str):
        try:
            dt = datetime.fromisoformat(tarih)  # '2024-05-19'
            return dt.strftime('%d.%m.%Y')
        except:
            return tarih
    return tarih

def recursive_format(value):
    if isinstance(value, list):
        return [recursive_format(v) for v in value]
    elif isinstance(value, dict):
        return {k: recursive_format(v) for k, v in value.items()}
    else:
        return tarih_formatla(value)

def render_template_tarihli(template_name, **context):
    formatted_context = {k: recursive_format(v) for k, v in context.items()}
    return render_template(template_name, **formatted_context)




@app.route('/login', methods=['GET', 'POST'])
def login():
    hata = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        conn = get_db_connection()
        user = conn.execute('SELECT * FROM kullanicilar WHERE username = ?', (username,)).fetchone()
        conn.close()

        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            flash('Başarıyla giriş yapıldı.')
            return redirect(url_for('admin_rehber'))
        else:
            hata = '❌ Kullanıcı adı veya şifre yanlış.'

    return render_template('login.html', hata=hata)

@app.route('/logout')
def logout():
    session.clear()
    flash('Başarıyla çıkış yapıldı.')
    return redirect(url_for('login'))

@app.route('/admin_rehber', methods=['GET'])
def admin_rehber():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    kisiler = conn.execute('SELECT * FROM kisiler').fetchall()
    conn.close()

    # Alfabetik olarak isme göre sırala
    kisiler = sorted(kisiler, key=lambda x: x['isim'].lower())

    return render_template_tarihli('panel_template/rehber_edit.html', rehber=kisiler)


@app.route('/ekle', methods=['POST'])
def ekle():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    if 'csv_file' in request.files:
        return redirect(url_for('ekle_csv'))  # Yeni route’a yönlendir

    isim = request.form['isim']
    departman = request.form['departman']
    dahili = request.form['dahili']
    email = request.form.get('email', '')
    cep = request.form.get('cep', '')

    conn = get_db_connection()
    conn.execute(
        'INSERT INTO kisiler (isim, departman, dahili, email, cep) VALUES (?, ?, ?, ?, ?)',
        (isim, departman, dahili, email, cep)
    )
    conn.commit()
    conn.close()

    flash('Yeni kişi başarıyla eklendi.')
    return redirect(url_for('admin_rehber'))


    # CSV içe aktarımı kontrolü
@app.route('/ekle_csv', methods=['POST'])
def ekle_csv():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    file = request.files.get('csv_file')
    if file and file.filename.endswith('.csv'):
        stream = StringIO(file.stream.read().decode("UTF8"), newline=None)
        reader = csv.DictReader(stream)
        conn = get_db_connection()
        for row in reader:
            conn.execute('''INSERT INTO kisiler (isim, departman, dahili, email, cep)
                            VALUES (?, ?, ?, ?, ?)''',
                         (row.get('isim'), row.get('departman'), row.get('dahili'), row.get('email'), row.get('cep')))
        conn.commit()
        conn.close()
        flash('CSV başarıyla içe aktarıldı.')
    else:
        flash('Lütfen geçerli bir CSV dosyası seçin.')

    return redirect(url_for('admin_rehber'))


@app.route('/guncelle', methods=['POST'])
def guncelle():
    # Güncelleme kodlarınız burada
    guncelle_id = request.form.get('guncelle_id')
    if not guncelle_id:
        flash('Güncellenecek kişi seçilmedi.')
        return redirect(url_for('admin_rehber'))

    isim = request.form.get(f'isim_{guncelle_id}')
    departman = request.form.get(f'departman_{guncelle_id}')
    dahili = request.form.get(f'dahili_{guncelle_id}')
    email = request.form.get(f'email_{guncelle_id}')
    cep = request.form.get(f'cep_{guncelle_id}')

    conn = get_db_connection()
    conn.execute('''
        UPDATE kisiler SET isim = ?, departman = ?, dahili = ?, email = ?, cep = ?
        WHERE id = ?
    ''', (isim, departman, dahili, email, cep, guncelle_id))
    conn.commit()
    conn.close()

    flash('Kişi bilgileri başarıyla güncellendi.')
    return redirect(url_for('admin_rehber'))

@app.route('/export_excel')
def export_excel():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    kisiler = conn.execute('SELECT isim, departman, dahili, email, cep FROM kisiler').fetchall()
    conn.close()

    kisiler_liste = [dict(kisi) for kisi in kisiler]

    df = pd.DataFrame(kisiler_liste, columns=['isim', 'departman', 'dahili', 'email', 'cep'])
    df.columns = ['İsim Soyisim', 'Departman', 'Dahili No', 'E-Posta', 'Cep Telefonu']

    # Yeni Excel dosyası
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Rehber'

    # Veri çerçeveye yaz
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Tüm hücreler ortalı
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Başlık satırına kalın yazı ve arka plan
            if r_idx == 1:
                cell.font = Font(bold=True)

            # Kenarlık ekle
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    # Sütunları otomatik genişlet
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    # Excel dosyasını belleğe yaz
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="telefon_rehberi.xlsx",
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/sil/<int:id>')
def sil(id):
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    conn.execute('DELETE FROM kisiler WHERE id = ?', (id,))
    conn.commit()
    conn.close()

    flash('Kişi başarıyla silindi.')
    return redirect(url_for('admin_rehber'))



@app.route('/rehber')
def rehber():
    conn = get_db_connection()
    kisiler = conn.execute('SELECT * FROM kisiler').fetchall()
    conn.close()

    rehber = sorted(kisiler, key=lambda x: x['isim'].lower())

    return render_template_tarihli('rehber.html', rehber=rehber)


@app.route('/admin/duyurular', methods=['GET', 'POST'])
def admin_duyurular():
    db = get_db_connection()

    # POST: Yeni duyuru ekleme
    if request.method == 'POST':
        baslik = request.form.get('baslik')
        icerik = request.form.get('icerik')
        tarih = request.form.get('tarih')  # opsiyonel, yoksa bugünün tarihi atanabilir

        if not baslik or not icerik:
            flash("Başlık ve içerik zorunludur.")
            return redirect(url_for('admin_duyurular'))

        if not tarih:
            from datetime import datetime
            tarih = datetime.now().strftime('%Y-%m-%d')

        db.execute(
            "INSERT INTO duyurular (baslik, icerik, tarih) VALUES (?, ?, ?)",
            (baslik, icerik, tarih)
        )
        db.commit()
        db.close()
        flash("Duyuru başarıyla eklendi.")
        
        return redirect(url_for('admin_duyurular'))

    # GET: Listeleme + filtreleme
    tarih_filter = request.args.get('tarih_filter')
    siralama = request.args.get('siralama', 'desc')

    query = "SELECT * FROM duyurular"
    params = []

    if tarih_filter:
        query += " WHERE tarih >= ?"
        params.append(tarih_filter)

    query += f" ORDER BY tarih {siralama.upper()}"

    duyurular = db.execute(query, params).fetchall()
    db.close()

    return render_template_tarihli('panel_template/duyuru_edit.html', duyurular=duyurular, tarih_filter=tarih_filter, siralama=siralama)

@app.route('/admin/duyuru_sil/<int:id>', methods=['GET'])
def admin_duyuru_sil(id):
    db = get_db_connection()
    db.execute("DELETE FROM duyurular WHERE id = ?", (id,))
    db.commit()
    db.close()
    flash("Duyuru başarıyla silindi.")
    return redirect(url_for('admin_duyurular'))


from datetime import datetime

@app.route('/duyurular')
def duyurular():
    conn = get_db_connection()
    duyurular_raw = conn.execute('SELECT * FROM duyurular ORDER BY tarih DESC').fetchall()
    conn.close()

    duyurular = []
    for d in duyurular_raw:
        tarih_iso = d['tarih']  # örn: 2024-05-20
        try:
            tarih_tr = datetime.strptime(tarih_iso, "%Y-%m-%d").strftime("%d.%m.%Y")
        except:
            tarih_tr = tarih_iso  # hata olursa olduğu gibi göster

        duyurular.append({
            'id': d['id'],
            'baslik': d['baslik'],
            'icerik': d['icerik'],
            'tarih': tarih_tr
        })

    return render_template('duyurular.html', duyurular=duyurular)


# Kullanıcı için yemek listesi sayfası (sadece görüntüleme)
@app.route('/yemek-listesi')
def yemek_listesi():
    conn = get_db_connection()
    yemekler = conn.execute('SELECT * FROM yemekler ORDER BY tarih').fetchall()
    conn.close()
    return render_template_tarihli('yemek_listesi.html', yemekler=yemekler)

# Admin için yemek listesi sayfası (düzenleme yapacak)
@app.route('/admin/yemek-listesi')
def admin_yemek_listesi():
    conn = get_db_connection()
    yemekler = conn.execute('SELECT * FROM yemekler ORDER BY tarih').fetchall()
    conn.close()
    return render_template_tarihli('panel_template/yemek_edit.html', yemekler=yemekler)

# Admin - yemek ekleme
@app.route('/admin/yemek-ekle', methods=['POST'])
def admin_yemek_ekle():
    tarih = request.form['tarih']
    yemek1 = request.form.get('yemek1', '')
    yemek2 = request.form.get('yemek2', '')
    yemek3 = request.form.get('yemek3', '')
    yemek4 = request.form.get('yemek4', '')
    yemek5 = request.form.get('yemek5', '')

    conn = get_db_connection()
    conn.execute('''
        INSERT INTO yemekler (tarih, yemek1, yemek2, yemek3, yemek4, yemek5) 
        VALUES (?, ?, ?, ?, ?, ?)''',
        (tarih, yemek1, yemek2, yemek3, yemek4, yemek5)
    )
    conn.commit()
    conn.close()

    flash('Yemek eklendi.')
    return redirect(url_for('admin_yemek_listesi'))



# Admin - yemek güncelleme
@app.route("/admin/yemek-guncelle", methods=["POST"])
def admin_yemek_guncelle():
    guncelle_id = request.form.get("guncelle_id")

    if not guncelle_id:
        flash("Güncellenecek öğe seçilemedi.")
        return redirect(url_for("admin_yemek_listesi"))

    tarih = request.form.get(f"tarih_{guncelle_id}")
    yemek1 = request.form.get(f"yemek1_{guncelle_id}")
    yemek2 = request.form.get(f"yemek2_{guncelle_id}")
    yemek3 = request.form.get(f"yemek3_{guncelle_id}")
    yemek4 = request.form.get(f"yemek4_{guncelle_id}")
    yemek5 = request.form.get(f"yemek5_{guncelle_id}")

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE yemekler
            SET tarih = ?, yemek1 = ?, yemek2 = ?, yemek3 = ?, yemek4 = ?, yemek5 = ?
            WHERE id = ?
        """, (tarih, yemek1, yemek2, yemek3, yemek4, yemek5, guncelle_id))
        conn.commit()
        conn.close()
        flash("Yemek bilgisi başarıyla güncellendi.")
    except Exception as e:
        flash(f"Güncelleme sırasında hata oluştu: {e}")

    return redirect(url_for("admin_yemek_listesi"))

    
# Admin - yemek silme
@app.route('/admin/yemek-sil/<int:yemek_id>')
def admin_yemek_sil(yemek_id):
    conn = get_db_connection()
    conn.execute('DELETE FROM yemekler WHERE id = ?', (yemek_id,))
    conn.commit()
    conn.close()
    flash('Yemek silindi.')
    return redirect(url_for('admin_yemek_listesi'))

# Admin - CSV import
@app.route('/export_yemek_listesi')
def export_yemek_listesi():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = get_db_connection()
    yemekler = conn.execute('SELECT tarih, menu FROM yemek_listesi ORDER BY tarih').fetchall()
    conn.close()

    # Verileri dönüştür
    yemek_liste = [dict(yemek) for yemek in yemekler]
    df = pd.DataFrame(yemek_liste, columns=['tarih', 'menu'])
    df.columns = ['Tarih', 'Menü']

    # Excel dosyası oluştur
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Yemek Listesi'

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Ortala
            cell.alignment = Alignment(horizontal='center', vertical='center')

            # Başlık kalın olsun
            if r_idx == 1:
                cell.font = Font(bold=True)

            # Kenarlık
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    # Sütunları genişlet
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="yemek_listesi.xlsx",
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/upload', methods=['GET', 'POST'])
def upload():
    conn = sqlite3.connect('rehber.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()

    c.execute('SELECT * FROM siparisler')
    siparisler = c.fetchall()
    siparisler_dict = [dict(row) for row in siparisler]

    conn.close()

    return render_template('panel_template/upload_siparis.html', siparisler=siparisler_dict)

@app.route('/siparisler')
def siparis_listesi():
    conn = sqlite3.connect('rehber.db')  # veritabanı dosyanız
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute('SELECT * FROM siparisler')
    rows = c.fetchall()
    kolonlar = [description[0] for description in c.description]  # sütun başlıkları
    conn.close()
    
    return render_template("siparis_listesi.html", siparisler=rows, kolonlar=kolonlar)

@app.route("/siparis_analiz")
def siparis_analiz():
    conn = get_db_connection()
    siparisler = conn.execute("SELECT * FROM siparisler").fetchall()
    conn.close()

    toplam_siparis = len(siparisler)

    gecikmeli_sayisi = 0
    termin_aylar = []
    terminine_yetisen = 0
    gecikme_sureleri = []
    musteri_gecikme_sayisi = defaultdict(int)

    now_date = datetime.now().date()

    for s in siparisler:
        termin = s["termin"]
        teslim = s["teslim_tarihi"] if "teslim_tarihi" in s.keys() else None
        musteri = s["musteri"]
        
        if termin:
            try:
                termin_tarihi = datetime.strptime(termin, "%Y-%m-%d %H:%M:%S").date()
                ay = termin_tarihi.strftime("%Y-%m")
                termin_aylar.append(ay)

                if termin_tarihi < now_date:
                    gecikmeli_sayisi += 1
                    musteri_gecikme_sayisi[musteri] += 1
                    
                    if teslim:
                        try:
                            teslim_tarihi = datetime.strptime(teslim, "%Y-%m-%d %H:%M:%S").date()
                            gecikme_sure = (teslim_tarihi - termin_tarihi).days
                            if gecikme_sure > 0:
                                gecikme_sureleri.append(gecikme_sure)
                        except:
                            pass

                else:
                    terminine_yetisen += 1

            except ValueError:
                pass

    termin_ay_sayim = dict(Counter(termin_aylar))

    # Ortalama gecikme süresi
    ort_gecikme_suresi = sum(gecikme_sureleri) / len(gecikme_sureleri) if gecikme_sureleri else 0

    oran_yetisen = (terminine_yetisen / toplam_siparis * 100) if toplam_siparis > 0 else 0
    oran_geciken = (gecikmeli_sayisi / toplam_siparis * 100) if toplam_siparis > 0 else 0

    musteri_sayim = Counter(s["musteri"] for s in siparisler if s["musteri"])
    en_cok_musteriler = musteri_sayim.most_common(5)
    en_cok_geciken_musteriler = sorted(musteri_gecikme_sayisi.items(), key=lambda x: x[1], reverse=True)[:5]

    # --- BURADA ORTALAMA HESABI VE AYLIK ORTALAMA DEĞİŞİMİ HESAPLAMA ---
    ay_sayisi = len(termin_ay_sayim) if termin_ay_sayim else 1
    ortalama_siparis = toplam_siparis / ay_sayisi

    aylik_ortalama_degisim = {}
    for ay, sayi in termin_ay_sayim.items():
        if ortalama_siparis > 0:
            degisim_orani = ((sayi - ortalama_siparis) / ortalama_siparis) * 100
        else:
            degisim_orani = 0
        aylik_ortalama_degisim[ay] = degisim_orani
    # --------------------------------------------------------------------

    return render_template("siparis_analiz.html",
                       toplam=toplam_siparis,
                       gecikme=gecikmeli_sayisi,
                       termin_ay_sayim=termin_ay_sayim,
                       oran_yetisen=oran_yetisen,
                       oran_geciken=oran_geciken,
                       en_cok_musteriler=en_cok_musteriler,
                       en_cok_geciken_musteriler=en_cok_geciken_musteriler,
                       ort_gecikme_suresi=ort_gecikme_suresi,
                       aylik_ortalama_degisim=aylik_ortalama_degisim)

@app.route('/export_siparisler')
def export_siparisler():
    conn = sqlite3.connect('rehber.db')
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute('SELECT * FROM siparisler')
    rows = c.fetchall()
    columns = [desc[0] for desc in c.description]
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Siparişler"

    # Stil tanımları
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")  # Mavi ton
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # Başlıkları yaz ve stil uygula
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    # Verileri yaz ve kenarlık ekle
    for row_num, row_data in enumerate(rows, 2):
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_num, value=row_data[column_title])
            cell.border = thin_border
            # Sayısal veriler için hizalama sağa olabilir, text için sola
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Sütun genişliklerini otomatik ayarlama (basit)
    for col_num, column_title in enumerate(columns, 1):
        max_length = len(column_title)
        for row_num in range(2, len(rows) + 2):
            cell = ws.cell(row=row_num, column=col_num)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = adjusted_width

    # Dosyayı belleğe kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, 
                    download_name="siparisler.xlsx", 
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)

